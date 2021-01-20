import pandas as pd
import tkinter as tk
from tkinter import Frame
from datetime import datetime
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from scipy.interpolate import interp1d
import scipy.optimize as optimize
from scipy.optimize import fsolve, root
import pathlib


def read_eis_data(file):
    # read out file data (experiement specs)
    # readout file data

    # read out experiment specifications
    df_eis_specs = pd.read_csv(file, decimal=',', encoding='cp1252', error_bad_lines=False, delim_whitespace=True, index_col=False,
                               header=None, nrows=7, keep_default_na=False)

    eis_specs_date = df_eis_specs[3].values[0]
    eis_specs_time = df_eis_specs[1].values[5]
    eis_specs_datetime = datetime.strptime(eis_specs_date + ' ' + eis_specs_time, '%b,%d.%Y %H:%M:%S')

    eis_specs_voltage = df_eis_specs[1].values[2][0:-1]
    eis_specs_file = df_eis_specs[1].values[0]
    eis_specs_system = df_eis_specs[1].values[1]

    # read and format measurement data !fix with load, save, reload, columns --> to get into aspired design
    df_eis_data = pd.read_csv(file, sep='\t', decimal='.', encoding='cp1252', error_bad_lines=False, skiprows=19, index_col=False)


    df_eis_data.to_csv('temp_eis_data.txt', mode='w', header=True)

    df_eis_data_formated = pd.read_csv('temp_eis_data.txt', decimal='.', encoding='cp1252', error_bad_lines=False,
                                       delim_whitespace=True)

    df_eis_data_formated.columns = ['index', 'number', 'frequency [Hz]', 'impedance Re [Ohm]', 'impedance Im [Ohm]',
                                    'significance', 'time [s]']

    df_eis_data_formated['Re [Ohm*cm²]'] = df_eis_data_formated['impedance Re [Ohm]'].apply(lambda x: x*25)
    df_eis_data_formated['-Im [Ohm*cm²]'] = df_eis_data_formated['impedance Im [Ohm]'].apply(lambda x: x*-25)



    return df_eis_data_formated, df_eis_specs, eis_specs_datetime, eis_specs_voltage, eis_specs_file, eis_specs_system


def read_qms_data(file):
    # read out file data (experiement specs)
    # readout file data

    # read out experiment specifications

    df_qms_specs = pd.read_csv(file, decimal=',', encoding='cp1252', error_bad_lines=False, delim_whitespace=True,
                               index_col=False, header=None, skiprows=2, nrows=5, keep_default_na=False)

    qms_filename = file.split('/')[-1][:-4]
    qms_specs_date = df_qms_specs[12].values[0]
    qms_specs_time = df_qms_specs[13].values[0]

    qms_specs_datetime = datetime.strptime(qms_specs_date + ' ' + qms_specs_time, '%Y-0%m-%d %H\'%M\'%S.isi')

    # read and format measurement data !fix with load, save, reload, columns --> to get into aspired design
    df_qms_data = pd.read_csv(file, sep='\t', decimal='.', encoding='cp1252', error_bad_lines=False, skiprows=7,
                              index_col=False)


    return df_qms_data, df_qms_specs, qms_specs_datetime, qms_filename


def save_eis_data(df_eis, filename, frame, name, molarity, current, voltage):

    df_eis['Sample'] = name
    df_eis['MeOH Molarity [M]'] = molarity
    df_eis['Current [mA]'] = current*1000
    df_eis.round(4)

    # save eis data to csv
    rootpath = pathlib.Path(__file__).parent.absolute()
    print(rootpath)
    df_eis.to_csv(str(rootpath)+'/EIS_data/CSV_data/'+filename+'.csv', mode='w', header=True,
                  index=False, sep='\t')

    # save eis data to excel
    wb_path = str(rootpath) + '/EIS_data/eis_data_library.xlsx'
    book = load_workbook(wb_path)
    writer = pd.ExcelWriter(wb_path, engine='openpyxl')
    writer.book = book
    df_eis.to_excel(writer, sheet_name=filename, index=False)

    writer.save()
    writer.close()

    #close frame
    frame.destroy()

    # preview
    x_values = np.asarray(df_eis['Re [Ohm*cm²]'])
    y_values = np.asarray(df_eis['-Im [Ohm*cm²]'])

    # graph interpolation
    f = interp1d(x_values, y_values, kind='linear', fill_value="extrapolate")

    roots = fsolve(lambda x: f(x), [-100, 100])
    hfr = roots[0]
    rr = roots[1] - roots[0]

    current = int(current) * 1000

    plt.plot(x_values, y_values, 'rs--')
    plt.plot(x_values, f(x_values), 'k-')
    plt.title('Preview - Nyquistplot (' + name + ' / ' + str(current) + 'mA / ' + molarity + 'M)', fontsize=16, pad=20)
    plt.xlabel('Re [Ohm*cm²]')
    plt.ylabel('-Im [Ohm*cm²]')
    plt.xticks()
    plt.axhline(0, color='black')

    res_table_data = [["High Frequency Resistance [Ohm*cm²]", str(round(hfr, 2))],
                  ["Reaction Resistance [Ohm*cm²]", str(round(rr, 2))]]
    res_table = plt.table(cellText=res_table_data,
                          bbox=[0, 0.85, 0.35, 0.1], colWidths=[.9, .1], cellLoc='right', edges='open')

    for (row, col), cell in res_table.get_celld().items():
        if col == 0:
            cell.set_height(1.3)
            cell._loc = 'left'
            cell.set_text_props(ma='left', fontweight=50)
        elif col == 1:
            cell.set_height(1.3)
            cell._loc = 'right'
            cell.set_text_props(ma='right')

    res_table.auto_set_font_size(False)
    res_table.set_fontsize(12)


    v_table_data = [['Cell Voltage [V]', str(voltage)[1:]]]
    v_table = plt.table(cellText=v_table_data,
                        bbox=[0.9, 0.95, 0.1, 0.05], colWidths=[.6, .4], cellLoc='right', edges='open')

    for (row, col), cell in v_table.get_celld().items():
        if col == 0:
            cell.set_height(1.3)
            cell._loc = 'left'
            cell.set_text_props(ma='left', fontweight=50)
        elif col == 1:
            cell.set_height(1.3)
            cell._loc = 'right'
            cell.set_text_props(ma='right')

    v_table.auto_set_font_size(False)
    v_table.set_fontsize(12)



    plt.show()


def save_qms_data(df_qms, filename, frame, name, temperature, molarity, current, bool_dmfc, bool_dmec):


    df_qms['Sample'] = name
    if bool_dmfc == 'DMFC':
        df_qms['System'] = 'DMFC'
    else:
        df_qms['System'] = 'DMEC'
    df_qms['Temperature'] = temperature
    df_qms['MeOH Molarity [M]'] = molarity
    df_qms['Current [mA]'] = current
    #df_qms.round(4)

    # save eis data to csv
    rootpath = pathlib.Path(__file__).parent.absolute()
    print(rootpath)
    df_qms.to_csv(str(rootpath)+'/QMS_data/CSV_data/' + filename + '.csv', mode='w', header=True, index=False, sep='\t')

    # save eis data to excel
    wb_path = str(rootpath) + '/QMS_data/qms_data_library.xlsx'
    book = load_workbook(wb_path)
    writer = pd.ExcelWriter(wb_path, engine='openpyxl')
    writer.book = book
    df_qms.to_excel(writer, sheet_name=filename, index=False)

    writer.save()
    writer.close()

    #close frame
    frame.destroy()

    # preview
    x_values = np.asarray(df_qms['Time Relative (sec)'])
    y_values_co2 = np.asarray(df_qms['Sig:_CO2_in_%'])
    y_values_o2 = np.asarray(df_qms['Sig:_O2_in_%'])
    y_values_n2 = np.asarray(df_qms['Sig:_N2_in_%'])
    y_values_h2 = np.asarray(df_qms['Sig:_H2_in_%'])
    y_values_ch3oh = np.asarray(df_qms['Sig:_MeOH_in_%'])

    plt.plot(x_values, y_values_co2, 'r-', label='CO2')
    plt.plot(x_values, y_values_o2, 'b-', label='O2')
    plt.plot(x_values, y_values_n2, 'g-', label='N2')
    plt.plot(x_values, y_values_h2, 'k-', label='H2')
    plt.plot(x_values, y_values_ch3oh, 'y-', label='MeOH')
    plt.xlim(0, 600)
    plt.yscale('log')
    plt.title('Preview - QMS (' + name + ' / ' + str(current) + 'mA / ' + molarity + 'M / ' + temperature + '°C)', fontsize=16, pad=20)
    plt.xlabel('time [s]')
    plt.ylabel('fraction [%]')
    plt.legend(loc='upper left')

    avg_co2_fraction = df_qms['Sig:_CO2_in_%'].mean()
    avg_n2_fraction = df_qms['Sig:_N2_in_%'].mean()
    avg_o2_fraction = df_qms['Sig:_O2_in_%'].mean()
    avg_h2_fraction = df_qms['Sig:_H2_in_%'].mean()
    avg_meoh_fraction = df_qms['Sig:_MeOH_in_%'].mean()


    fraction_table_data = [['N2', str(round(avg_n2_fraction, 2))], ['O2', str(round(avg_o2_fraction, 2))],
                           ['H2', str(round(avg_h2_fraction, 2))], ['CO2', str(round(avg_co2_fraction, 2))],
                           ['MeOH', str(round(avg_meoh_fraction, 2))]
                           ]

    fraction_table = plt.table(cellText=fraction_table_data,
                          bbox=[0.78, 0.5, 0.2, 0.2], colWidths=[.6, .4], colLabels=['Avg. mass fraction', '[%]'])

    for (row, col), cell in fraction_table.get_celld().items():
        if col == 0:
            cell.set_height(1.3)
            cell._loc = 'left'
            cell.set_text_props(ma='left', fontweight=50)
        elif col == 1:
            cell.set_height(1.3)
            cell._loc = 'right'
            cell.set_text_props(ma='right')
        if row == 4:
            cell.set_text_props(color='r')

    # fraction_table.auto_set_font_size(False)
    # fraction_table.set_fontsize(10)


    plt.show()


def import_eis_data(frame, file):


    # implement sample_documentation_frame
    sampledoc_frame = tk.Toplevel(frame)
    sampledoc_frame.title("Additional Sample Data")
    sampledoc_frame.geometry("{}x{}".format(400, 400))
    sampledoc_frame.maxsize(400, 400)
    sampledoc_frame.config(bg="blue")
    sampledoc_frame.iconbitmap('zbt_logo.ico')


    # subframes of sampledoc_frame
    top_sampledoc_frame = Frame(sampledoc_frame, bg='lightgrey', width=400,
                                height=300)
    bot_sampledoc_frame = Frame(sampledoc_frame, bg='grey', width=400,
                                height=100)
    top_sampledoc_frame.grid_propagate(0)
    bot_sampledoc_frame.grid_propagate(0)

    top_sampledoc_frame.grid(row=0)
    bot_sampledoc_frame.grid(row=1)

    # subsubframes
    # top_left_sampledoc_frame = Frame(top_sampledoc_frame, bg='lightblue', width=300, height=350)
    # top_right_sampledoc_frame = Frame(top_sampledoc_frame, bg='lightgreen', width=500, height=350)
    # top_left_sampledoc_frame.grid_propagate(0)
    # top_right_sampledoc_frame.grid_propagate(0)

    # top_left_sampledoc_frame.grid(row=0, column=0)
    # top_right_sampledoc_frame.grid(row=0, column=1)

    df_eis, df_eis_specs, eis_datetime, eis_voltage, eis_file, eis_system = read_eis_data(file)

    # labels of entries
        # labels (date, sample ,molarity, current, voltage)
    sdf_label1 = tk.Label(top_sampledoc_frame, text='Date:', pady=10,
                          bg='lightgrey')
    sdf_label2 = tk.Label(top_sampledoc_frame, text='Sample:', pady=10,
                          bg='lightgrey')
    sdf_label3 = tk.Label(top_sampledoc_frame, text='Molarity [M]:', pady=10,
                          bg='lightgrey')
    sdf_label4 = tk.Label(top_sampledoc_frame, text='Current [A]:', pady=10,
                          bg='lightgrey')
    sdf_label5 = tk.Label(top_sampledoc_frame, text='Voltage [V]:', pady=10,
                          bg='lightgrey')

    CheckVar1 = IntVar()
    CheckVar2 = IntVar()

    sdf_cb1 = tk.Checkbutton(top_sampledoc_frame, text="DMFC", variable=CheckVar1,
                     onvalue=1, offvalue=0, height=5, width=20)

    sdf_cb2 = tk.Checkbutton(top_sampledoc_frame, text="DMEC", variable=CheckVar2,
                     onvalue=1, offvalue=0, height=5, width=20)



    sdf_label7 = tk.Label(bot_sampledoc_frame, text='File:   ' + '\t' + eis_file,
                          pady=10, bg='grey')
    sdf_label8 = tk.Label(bot_sampledoc_frame, text='System: ' + '\t' +
                                                    eis_system,
                          pady=10, bg='grey')

    # label placement
    sdf_label1.grid(row=1, column=0, sticky='w', padx=(10, 10))
    sdf_label2.grid(row=2, column=0, sticky='w', padx=(10, 10))
    sdf_label3.grid(row=4, column=0, sticky='w', padx=(10, 10))
    sdf_label4.grid(row=5, column=0, sticky='w', padx=(10, 10))
    sdf_label5.grid(row=6, column=0, sticky='w', padx=(10, 10))

    sdf_label7.grid(row=0, column=0, sticky='w', padx=(10, 0))
    sdf_label8.grid(row=1, column=0, sticky='w', padx=(10, 0))

    # cb placement
    sdf_cb1.grid(row=3, column=0, sticky='w', padx=(10, 10))
    sdf_cb2.grid(row=3, column=0, sticky='w', padx=(10, 10))

    # entries for additonal documentation

        # entries (date, sample ,molarity, current, voltage)
    sdf_entry1 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry1.insert(0, eis_datetime)
    sdf_entry2 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry2.insert(0, 'xxxx')
    sdf_entry3 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry3.insert(0, '0')
    sdf_entry4 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry4.insert(0, '0')
    sdf_entry5 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry5.insert(0, eis_voltage)

    # entry placement
    sdf_entry1.grid(row=1, column=1)
    sdf_entry2.grid(row=2, column=1)
    sdf_entry3.grid(row=3, column=1)
    sdf_entry4.grid(row=4, column=1)
    sdf_entry5.grid(row=5, column=1)

    # buttons of sampledoc_frame
        # sdf_button1 --> format file and store with given additional data
    sdf_button1 = tk.Button(top_sampledoc_frame, text='OK', width=20, bd=5,
                            command=lambda: save_eis_data(df_eis, eis_file,
                                                          sampledoc_frame,
                                                          sdf_entry2.get(),
                                                          sdf_entry3.get(),
                                                          sdf_entry4.get(),
                                                          sdf_entry5.get()))
    sdf_button1.grid(row=6, column=1)


    sampledoc_frame.mainloop()


def import_qms_data(frame, file):

    # implement sample_documentation_frame
    sampledoc_frame = tk.Toplevel(frame)
    sampledoc_frame.title("Additional Sample Data")
    sampledoc_frame.geometry("{}x{}".format(400, 300))
    sampledoc_frame.maxsize(400, 300)
    sampledoc_frame.config(bg="blue")
    sampledoc_frame.iconbitmap('zbt_logo.ico')

    # subframes of sampledoc_frame
    top_sampledoc_frame = Frame(sampledoc_frame, bg='lightgrey', width=400, height=250)
    bot_sampledoc_frame = Frame(sampledoc_frame, bg='grey', width=400, height=50)
    top_sampledoc_frame.grid_propagate(0)
    bot_sampledoc_frame.grid_propagate(0)

    top_sampledoc_frame.grid(row=0)
    bot_sampledoc_frame.grid(row=1)

    df_qms, df_qms_specs, qms_datetime, qms_file = read_qms_data(file)

    # labels of entries
        # labels (date, sample ,molarity, current, voltage)
    sdf_label1 = tk.Label(top_sampledoc_frame, text='Date:', pady=10, bg='lightgrey')
    sdf_label2 = tk.Label(top_sampledoc_frame, text='Sample:', pady=10, bg='lightgrey')
    sdf_label3 = tk.Label(top_sampledoc_frame, text='Temperature [°C]:', pady=10, bg='lightgrey')
    sdf_label4 = tk.Label(top_sampledoc_frame, text='Molarity [M]:', pady=10, bg='lightgrey')
    sdf_label5 = tk.Label(top_sampledoc_frame, text='Current [mA]:', pady=10, bg='lightgrey')

    sdf_label7 = tk.Label(bot_sampledoc_frame, text='File: ' + qms_file, pady=10, bg='grey')

    # checkbuttons
    # checkbutton_choices (dmfc, dmec)
    cb1_var ='1'
    cb2_var = '2'

    sdf_cb1 = tk.Checkbutton(top_sampledoc_frame, text="DMFC",
                             variable=cb1_var,
                             onvalue='DMFC', offvalue=0)

    sdf_cb2 = tk.Checkbutton(top_sampledoc_frame, text="DMEC",
                             variable=cb2_var,
                             onvalue='DMEC', offvalue=0)

        # label placement
    sdf_label1.grid(row=1, column=0, sticky='ew', padx=(10, 10))
    sdf_label2.grid(row=2, column=0, sticky='ew', padx=(10, 10))
    sdf_label3.grid(row=4, column=0, sticky='ew', padx=(10, 10))
    sdf_label4.grid(row=5, column=0, sticky='ew', padx=(10, 10))
    sdf_label5.grid(row=6, column=0, sticky='ew')

    sdf_label7.grid(row=0, column=0, sticky='ew', padx=(10, 0))

    # cb placement
    sdf_cb1.grid(row=3, column=0, sticky='w', padx=(10, 10))
    sdf_cb2.grid(row=3, column=1, sticky='w', padx=(10, 10))

    # entries for additonal documentation

        # entries (date, sample ,molarity, current, voltage)
    sdf_entry1 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry1.insert(0, qms_datetime)
    sdf_entry2 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry2.insert(0, 'xxxx')
    sdf_entry3 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry3.insert(0, '0')
    sdf_entry4 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry4.insert(0, '0')
    sdf_entry5 = tk.Entry(top_sampledoc_frame, width=30, bd=5)
    sdf_entry5.insert(0, '0')

        # entry placement
    sdf_entry1.grid(row=1, column=1)
    sdf_entry2.grid(row=2, column=1)
    sdf_entry3.grid(row=4, column=1)
    sdf_entry4.grid(row=5, column=1)
    sdf_entry5.grid(row=6, column=1)

    # buttons of sampledoc_frame
        # sdf_button1 --> format file and store with given additional data
    sdf_button1 = tk.Button(top_sampledoc_frame, text='OK', width=20, bd=5,
                            command=lambda: save_qms_data(df_qms, qms_file,
                                                          sampledoc_frame,
                                                          sdf_entry2.get(),
                                                          sdf_entry3.get(),
                                                          sdf_entry4.get(),
                                                          sdf_entry5.get(),
                                                          sdf_cb1.getvar(cb1_var),
                                                          sdf_cb2.getvar(cb2_var)))
    sdf_button1.grid(row=7, column=1)


    sampledoc_frame.mainloop()




